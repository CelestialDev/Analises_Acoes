from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementClickInterceptedException, NoSuchElementException
from time import sleep
from playsound import playsound
import pyautogui as py
import sys
import os
from datetime import date
import pandas as pd
from IPython.display import display
from pandas import read_excel
from openpyxl import Workbook
#from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class WebScraping():
    def __init__(self):
        self.lista_nome_investimento = ["investimentos_especificos","investimentos_ibovespa","investimentos_financeiro",
        "investimentos_consumo","investimentos_geral","investimentos_imobiliario","investimentos_industrial","investimento_valor"]
        self.index_planilha = 2
        self.listas_acao = [] 
        self.lista_empresa = []
        self.lista_acao = []
        self.lista_precoatual = []
        self.lista_roe = []
        self.lista_roa = []
        self.lista_roic = []
        self.lista_pl = []
        self.lista_ebitda = []
        self.lista_ebit = []
        self.lista_liq_corrente = []
        self.lista_valor_crescimento_dia = []
        self.lista_valor_crescimento_mes = []
        self.lista_valor_crescimento_ano = []
        self.bdr = []
        self.aux_num = 0

    def Iniciar(self, aba, index_base, busca):
        self.navegador = webdriver.Chrome()
        os.system('cls' if os.name == 'nt' else 'clear')
        self.aba = aba
        self.index_base = index_base
        if busca != 1:
            self.acessar_site(self.aba, self.index_base)
        else:
            print("")
    
    def acessar_site(self, index, index_base):
        self.navegador.get('https://statusinvest.com.br/acoes/variacao/ibovespa')
        if self.aba == 1:
            self.buscar_acao()
        self.guardar_nome_acoes()

    def destruir_programa(self, index):
        playsound('error.wav')
        erro_destruir = input(f"\033[91mErro {index} ocorreu, tente reiniciar o programa.\033[m")
        sys.exit()

    def fechar_propaganda_padrao(self):
        try:
            self.btn1 = self.navegador.find_element_by_xpath('/html/body/div[12]/div/div/div[1]/button')
        finally:
            try:
                self.btn1.click()
                print(f'\033[0;34mRetirando propaganda padrão...\033[m') 
            except:
                self.destruir_programa("002")

    def fechar_propaganda_down(self):
        try:
            self.btn2 = self.navegador.find_element_by_xpath('/html/body/div[1]/div[2]')
        finally:
            try:
                self.btn2.click()
                print(f'\033[0;34mRetirando propaganda...\033[m') 
            except:
                self.destruir_programa("004")


    def buscar_acao(self): # Entrar na página da ação
        while True:
            self.search = input('Insira a ação que deseja buscar: ').upper()
            sleep(2)
            self.fechar_propaganda_padrao()
            try:
                self.navegador.find_element_by_xpath('//*[@id="main-nav-nav"]/div/div/div/ul/li[2]/a/i').click() # Aperta na lupa
            finally:
                self.barra = self.navegador.find_element_by_xpath('//*[@id="main-search"]/div[1]/span[1]/input[2]') # Localiza e manda "search" na barra
                self.barra.send_keys(self.search)
                try:
                    sleep(2)
                    self.navegador.find_element_by_xpath('//*[@id="main-search"]/div[2]/div/div/a/div/div[2]/div').click() # Clica na ação
                    break
                except:
                    self.destruir_programa("008")
                  
    def guardar_nome_acoes(self): # Guarda o nome das ações em uma lista 
        sleep(4) # Dorme para aparecer propaganda

        if self.aba == 1: # Aba pesquisa
            self.listas_acao.append(self.search)
            self.bdr.append(self.search)
                  
        elif self.aba == 2: # Aba investimentos alto e baixo
            self.trocar_filtro()
            print(f'\033[0;34mTrocando filtro...\033[m')   
            sleep(5) # Dorme pra trocar filtro
            
            try:
                btn_1 = self.navegador.find_element_by_xpath('/html/body/main/section[1]/div/div[2]/div[2]/div/input')
                btn_2 = self.navegador.find_element_by_xpath('/html/body/main/section[1]/div/div[2]/div[2]/div/ul/li[3]')
            finally:
                try:
                    btn_1.click() #Clicar em 10
                    btn_2.click() # Mudar para TODOS
                except:
                    self.destruir_programa("016")
            sleep(2) # Dorme para carregar Todos
            
            os.system('cls' if os.name == 'nt' else 'clear')
            print("\033[0;33mObtendo o nome das ações...\033[m")

            nome_acao_alta = self.navegador.find_elements_by_xpath('//span[@title="ticker/código do ativo"]')
            
            for item1 in nome_acao_alta:
                if item1.text != "":
                    self.listas_acao.append(item1.text)
                    self.bdr.append(item1)

            print("\033[91mAviso:\033[m O programa precisa que você esteja com a página do navegador "
            "aberta e no meio da tela do computador")
            sleep(10)
            py.scroll(10000000, x=531, y=553)
            print("\033[0;33mSubindo a página...\033[m")
            sleep(2)
            try:
                self.navegador.find_element_by_xpath("/html/body/main/section[1]/div/div[1]/ul/li[2]").click()
            except ElementClickInterceptedException:
                self.destruir_programa("032")

            sleep(2)
            nome_acao_baixa = self.navegador.find_elements_by_xpath('//span[@title="ticker/código do ativo"]')

            for item2 in nome_acao_baixa:
                if item2.text != "":
                    self.listas_acao.append(item2.text)
                    self.bdr.append(item2)
            os.system('cls' if os.name == 'nt' else 'clear')

        self.analisar_acao(0)
    
    def trocar_filtro(self):
        sleep(2)
        self.fechar_propaganda_padrao()
        sleep(4)
        try:
            self.fechar_propaganda_down()
        except:
            self.destruir_programa("064")
        finally:
            sleep(1)
            self.navegador.find_element_by_xpath('//*[@id="form-search"]/div/div[1]/div/input').click() # Clica no Ibovespa
        
        sleep(1)
        if self.index_base == 2: # Funcional 
            finan_btn = self.navegador.find_element_by_xpath("/html/body/main/section[1]/div/form/div/div[1]/div/ul/li[11]/span") # Clica no Indíce Financeiro
            finan_btn.click()
        elif self.index_base == 3: # Funcional 
            cons_btn = self.navegador.find_element_by_xpath("/html/body/main/section[1]/div/form/div/div[1]/div/ul/li[9]/span") # Clica no Indíce de Consumo
            cons_btn.click()
        elif self.index_base == 4: # Funcional 
            geral_btn = self.navegador.find_element_by_xpath("/html/body/main/section[1]/div/form/div/div[1]/div/ul/li[1]/span") # Clica no Indíce GERAL
            geral_btn.click()
        elif self.index_base == 5: # Funcional 
            imob_btn = self.navegador.find_element_by_xpath("/html/body/main/section[1]/div/form/div/div[1]/div/ul/li[16]/span") # Clica no Indíce Imobiliário
            imob_btn.click()
        elif self.index_base == 6: # Funcional 
            indus_btn = self.navegador.find_element_by_xpath("/html/body/main/section[1]/div/form/div/div[1]/div/ul/li[17]/span") # Clica no Indíce Industrial
            indus_btn.click()
        elif self.index_base == 7: # Funcional 
            sleep(2)
            val_btn = self.navegador.find_element_by_xpath("/html/body/main/section[1]/div/form/div/div[1]/div/ul/li[20]/span") # Clica no Indíce Valor
            val_btn.click()

    def analisar_acao(self, index): # Obtenção dos valores de cada ação na lista
        # Obtenção de dados de cada ação
        for acao, bdr in zip(self.listas_acao, self.bdr):
            print(f'Dados carregados da ação \033[1;33m{acao}\033[m')
            print(f'\033[0;34mTranscrevendo dados para a planilha\033[m')   
            ros = []
            div = []

            if index == 0:      
                self.navegador.get(f'https://statusinvest.com.br/acoes/{acao}')
            elif index == 1:
                self.navegador.get(f'https://statusinvest.com.br/{bdr}/{acao}')

            # Informação do nome da empresa
            empresa = self.navegador.find_element_by_xpath('/html/body/main/header/div/div/div[1]/h1/small')

            # Informação do preço atual
            precoatual = self.navegador.find_element_by_xpath('//*[@id="main-2"]'
            '/div[2]/div/div[1]/div/div[1]/div/div[1]/strong')
   
            preco = precoatual.text
            preco_tratado = preco.replace(',', '.')
            self.lista_empresa.append(empresa.text)
            self.lista_acao.append(acao)
            self.lista_precoatual.append(preco_tratado)

            if index == 0:
                # Informação do ROE, ROA, ROIC
                for x in range(3):
                    ros_aux = self.navegador.find_element_by_xpath(f'//*[@id="indicators-section"]/div[2]/div/div[4]/div/div[{x + 1}]/div/div/strong')

                    ros.append(ros_aux)
                        

                # Informação das dívidas: líquida/pl, líquida/ebitda, líquida/ebit, liq.corrente
                for x in range(3):
                    div_aux = self.navegador.find_element_by_xpath(f'//*[@id="indicators-section"]'
                    f'/div[2]/div/div[2]/div/div[{x + 1}]/div/div/strong')

                    div.append(div_aux)
            
                liq_corrente = self.navegador.find_element_by_xpath('//*[@id="indicators-section"]/div[2]/div/div[2]/div/div[6]/div/div/strong')
                liq_texto = liq_corrente.text
                liq_tratado = liq_texto.replace(',','.')

                # Informação se a ação está em alta ou em baixa DIA
                valor_crescimento = self.navegador.find_element_by_xpath('//*[@id="main-2"]'
                '/div[2]/div/div[1]/div/div[1]/div/div[2]/span/b')
            
                value_dia = valor_crescimento.text
                value_dia_tratado1 = value_dia.replace(',','.')
                value_dia_tratado = value_dia_tratado1.replace('%','')

                # Informação se a ação está em alta ou em baixa MES
                valor_crescimento_mes = self.navegador.find_element_by_xpath('//*[@id="main-2"]/div[2]/div/div[1]/div/div[5]/div/div[2]/div/span[2]/b')
            
                value_mes = valor_crescimento_mes.text
                value_mes_tratado1 = value_mes.replace(',','.')
                value_mes_tratado = value_mes_tratado1.replace('%','')

                # Informação se a ação está em alta ou em baixa ANO
                valor_crescimento_ano = self.navegador.find_element_by_xpath('//*[@id="main-2"]/div[2]/div/div[1]/div/div[5]/div/div[1]/strong')
            
                value_ano = valor_crescimento_ano.text
                value_ano_tratado1 = value_ano.replace(',','.')
                value_ano_tratado = value_ano_tratado1.replace('%','')
            
                # Variáveis obtidas nesta função: precoatual, ros, div, crescimento, valor_crescimento

                # Divisão das pastas a cada informação
                roe = ros[0]
                roe_texto = roe.text
                roe_tratado1 = roe_texto.replace(',','.')
                roe_tratado = roe_tratado1.replace('%','')

                roa = ros[1]
                roa_texto = roa.text
                roa_tratado1 = roa_texto.replace(',','.')
                roa_tratado = roa_tratado1.replace('%','')

                roic = ros[2]
                roic_texto = roic.text
                roic_tratado1 = roic_texto.replace(',','.')
                roic_tratado = roic_tratado1.replace('%','')

                pl = div[0]
                pl_texto = pl.text
                pl_tratado = pl_texto.replace(',','.')

                ebitda = div[1]
                ebitda_texto = ebitda.text
                ebitda_tratado = ebitda_texto.replace(',','.')

                ebit = div[2]
                ebit_texto = ebit.text
                ebit_tratado = ebit_texto.replace(',','.')

                # XPath para string
                self.lista_roe.append(roe_tratado)
                self.lista_roa.append(roa_tratado)
                self.lista_roic.append(roic_tratado)
                self.lista_pl.append(pl_tratado)
                self.lista_ebitda.append(ebitda_tratado)
                self.lista_ebit.append(ebit_tratado)
                self.lista_liq_corrente.append(liq_tratado)
                self.lista_valor_crescimento_dia.append(value_dia_tratado)
                self.lista_valor_crescimento_mes.append(value_mes_tratado)
                self.lista_valor_crescimento_ano.append(value_ano_tratado)
        playsound('conclued.wav')

        if self.aux_num == 0:   
            self.cria_planilha()
                  
    def cria_planilha(self):
        # Cabeçalho
        self.planilha = Workbook()
        self.planilha_investimentos = self.planilha['Sheet']
        self.planilha_investimentos.cell(row=1, column=1, value='EMPRESA')
        self.planilha_investimentos.cell(row=1, column=2, value='AÇÃO')
        self.planilha_investimentos.cell(row=1, column=3, value='PREÇO ATUAL')
        self.planilha_investimentos.cell(row=1, column=4, value='ROE')
        self.planilha_investimentos.cell(row=1, column=5, value='ROA')
        self.planilha_investimentos.cell(row=1, column=6, value='ROIC')
        self.planilha_investimentos.cell(row=1, column=7, value='DIV/PL')
        self.planilha_investimentos.cell(row=1, column=8, value='DIV/EBITDA')
        self.planilha_investimentos.cell(row=1, column=9, value='DIV/EBIT')
        self.planilha_investimentos.cell(row=1, column=10, value='LIQ.CORRENTE')
        self.planilha_investimentos.cell(row=1, column=11, value='VALORIZAÇÃO ATUAL')
        self.planilha_investimentos.cell(row=1, column=12, value='VALORIZAÇÃO MES')
        self.planilha_investimentos.cell(row=1, column=13, value='VALORIZAÇÃO ANO')

        # Dados , roe, roa, roic, pl, ebitda, ebit, crescimento, valor_crescimento
            
        for empresa, acao, preco, roe, roa, roic, pl, ebitda, ebit, liq_corrente, valor_crescimento_dia, valor_crescimento_mes, valor_crescimento_ano in zip(
            self.lista_empresa, self.lista_acao, self.lista_precoatual,self.lista_roe, self.lista_roa, self.lista_roic, self.lista_pl,
            self.lista_ebitda, self.lista_ebit, self.lista_liq_corrente, self.lista_valor_crescimento_dia, self.lista_valor_crescimento_mes,
            self.lista_valor_crescimento_ano):
            
            self.planilha_investimentos.cell(row= self.index_planilha, column=1, value=empresa)
            self.planilha_investimentos.cell(row= self.index_planilha, column=2, value=acao)
            self.planilha_investimentos.cell(row= self.index_planilha, column=3, value=preco)
            self.planilha_investimentos.cell(row= self.index_planilha, column=4, value=roe)
            self.planilha_investimentos.cell(row= self.index_planilha, column=5, value=roa)
            self.planilha_investimentos.cell(row= self.index_planilha, column=6, value=roic)
            self.planilha_investimentos.cell(row= self.index_planilha, column=7, value=pl)
            self.planilha_investimentos.cell(row= self.index_planilha, column=8, value=ebitda)
            self.planilha_investimentos.cell(row= self.index_planilha, column=9, value=ebit)
            self.planilha_investimentos.cell(row= self.index_planilha, column=10, value=liq_corrente)
            self.planilha_investimentos.cell(row= self.index_planilha, column=11, value=valor_crescimento_dia)
            self.planilha_investimentos.cell(row= self.index_planilha, column=12, value=valor_crescimento_mes)
            self.planilha_investimentos.cell(row= self.index_planilha, column=13, value=valor_crescimento_ano)
            self.index_planilha += 1
        
        for x in range(0,7):
            if self.index_base == x:
                nome_arquivo = self.lista_nome_investimento[x]
            
        self.planilha.save(f'pastaplanilhas\{nome_arquivo}.xlsx')   
        print(f'\033[0;32mTranscrição finalizada!\033[m')
        print(f'O arquivo \033[0;33m{nome_arquivo}\033[m está localizado na pasta "pastaplanilhas", na mesma pasta que o programa')       
        self.navegador.quit() 
    
    def acao_obtida(self, lista_dados):
        self.navegador = webdriver.Chrome()
        os.system('cls' if os.name == 'nt' else 'clear')
        self.aux_num = 1
        self.aux_item_num = 0
        self.empresa = []
        self.acao = []
        self.qtd = []
        self.valor = []
        self.bdr = []

        for x in range(len(lista_dados)):
            empresa = lista_dados[x][0]
            acao = lista_dados[x][1]
            qtd = lista_dados[x][2]
            valor_obtido = lista_dados[x][3]
            bdr = lista_dados[x][4]
            self.empresa.append(empresa)
            self.acao.append(acao)
            self.listas_acao.append(acao)
            self.qtd.append(qtd)
            self.bdr.append(bdr)
            self.valor.append(valor_obtido)

        self.planilha = Workbook()
        try:
            self.planilha = load_workbook('Acoes.xlsx')
        except:
            self.planilha_acoes = self.planilha['Sheet']
            self.planilha_acoes.cell(row=1, column=1, value='EMPRESA')
            self.planilha_acoes.cell(row=1, column=2, value='AÇÃO')
            self.planilha_acoes.cell(row=1, column=3, value='QUANTIDADE')
            self.planilha_acoes.cell(row=1, column=4, value='VALOR OBTIDO')

            for i1,i2,i3,i4 in zip(self.empresa, self.acao, self.qtd, self.valor):
                self.planilha_acoes.cell(row= self.index_planilha, column=1, value=i1)
                self.planilha_acoes.cell(row= self.index_planilha, column=2, value=i2)
                self.planilha_acoes.cell(row= self.index_planilha, column=3, value=i3)
                self.planilha_acoes.cell(row= self.index_planilha, column=4, value=i4)
                self.index_planilha += 1
            self.planilha.save('pastaplanilhas\Acoes.xlsx')
        
        self.analisar_acao(1)
        self.navegador.quit()
        acao_df = pd.read_excel('pastaplanilhas\Acoes.xlsx')
        data_atual = date.today()
        data_texto = '{}/{}/{}'.format(data_atual.day, data_atual.month,data_atual.year)
        for x,y in zip(self.qtd, self.lista_precoatual):
            acao_df.loc[self.aux_item_num,f'{data_texto}'] = int(x) * float(y)
            self.aux_item_num += 1
        display(acao_df)
        acao_df.to_excel('pastaplanilhas\Acoes.xlsx', index = False)